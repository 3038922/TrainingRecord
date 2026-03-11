import os
import json
import requests
from datetime import datetime, date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ======================
# 1) 配置
# ======================
TARGET_YEAR = 2025
folder_path = os.getcwd()

CONFIG = {
    "summer_time_start": 501,      # 夏令时开始日期（5月1日）
    "winter_break": (120, 225),    # 寒假（MMDD）
    "summer_break": (701, 831),    # 暑假（MMDD）
    "duration": {
        "summer": 0.5,
        "winter": 1.0,
        "friday_work_day": 3,  # 周五额外加课
        "holiday": 8.0
    }
}

# ======================
# 2) 节假日：联网 + 缓存（同时处理补班 workday）
# ======================
def get_cached_or_fetch_holidays_and_workdays(year: int):
    filename = f"holidays_{year}.json"

    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding="utf-8") as f:
                obj = json.load(f)
            return set(obj.get("holidays", [])), set(obj.get("workdays", []))
        except Exception as e:
            print(f"⚠️ 本地缓存读取失败：{e}")

    try:
        print(f"🌐 正在联网获取 {year} 年节假日数据…")
        url = f"https://timor.tech/api/holiday/year/{year}/"
        res = requests.get(url, timeout=8)
        res.raise_for_status()
        data = res.json()

        holidays = set()
        workdays = set()
        for day, info in data.get("holiday", {}).items():
            if not info:
                continue
            if info.get("holiday", False):
                holidays.add(day)
            if info.get("workday", False):
                workdays.add(day)

        with open(filename, "w", encoding="utf-8") as f:
            json.dump(
                {"holidays": sorted(holidays), "workdays": sorted(workdays)},
                f, ensure_ascii=False, indent=2
            )

        print(f"✅ 节假日数据已缓存到：{filename}")
        return holidays, workdays

    except Exception as e:
        print(f"⚠️ 获取节假日失败：{e}")
        print("🔁 使用兜底（仅元旦），其余按寒暑假等规则判断")
        holidays = {f"{year}-01-01"}
        workdays = set()

        with open(filename, "w", encoding="utf-8") as f:
            json.dump(
                {"holidays": sorted(holidays), "workdays": sorted(workdays)},
                f, ensure_ascii=False, indent=2
            )
        return holidays, workdays

# ======================
# 3) 判断逻辑
# ======================
def is_in_break(d: date) -> bool:
    md = d.month * 100 + d.day
    wb_start, wb_end = CONFIG["winter_break"]
    sb_start, sb_end = CONFIG["summer_break"]
    return (wb_start <= md <= wb_end) or (sb_start <= md <= sb_end)

def is_summer_time(d: date) -> bool:
    md = d.month * 100 + d.day
    return md >= CONFIG["summer_time_start"]

def is_legal_holiday(date_str: str, holiday_set: set, workday_set: set) -> bool:
    # ✅ 补班日优先：即使是周六也算工作日
    if date_str in workday_set:
        return False

    # ✅ 法定节假日
    if date_str in holiday_set:
        return True

    d = datetime.strptime(date_str, "%Y-%m-%d").date()

    # ✅ 周六也算假日（周日不生成，所以这里只管周六）
    if d.weekday() == 5:  # 5 = Saturday
        return True

    # ✅ 寒暑假也算假日
    if is_in_break(d):
        return True

    return False

def get_training_duration(d: date, is_holiday: bool) -> float:
    if is_holiday:
        return CONFIG["duration"]["holiday"]

    duration = CONFIG["duration"]["summer"] if is_summer_time(d) else CONFIG["duration"]["winter"]
    if d.weekday() == 4:  # Friday
        duration += CONFIG["duration"]["friday_work_day"]
    return duration

# ======================
# 4) Excel 样式
# ======================
def autofit_column_width(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

def style_header(sheet):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# ======================
# 5) 生成全年日期（周一~周六，跳过周日）
# ======================
holidays, workdays = get_cached_or_fetch_holidays_and_workdays(TARGET_YEAR)

month_data = {m: [] for m in range(1, 13)}
d = date(TARGET_YEAR, 1, 1)
end = date(TARGET_YEAR, 12, 31)

while d <= end:
    if d.weekday() != 6:  # 6=Sunday
        month_data[d.month].append(d.strftime("%Y-%m-%d"))
    d += timedelta(days=1)

# ======================
# 6) 写入 Excel（封面 + 1~12月，含“金额”列）
# ======================
wb = Workbook()
cover = wb.active
cover.title = "封面"

cover["A1"] = "月份"
cover["B1"] = "工作日训练时长"
cover["C1"] = "节假日训练时长"
cover["D1"] = "工作日金额（x100）"
cover["E1"] = "节假日金额（x200）"
cover["F1"] = "总金额"

for month in range(1, 13):
    sheet = wb.create_sheet(f"{month}月")

    # ✅ 增加金额列
    sheet["A1"] = f"{month}月训练日期"
    sheet["B1"] = "是否法定节假日"
    sheet["C1"] = "训练时长（小时）"
    sheet["D1"] = "金额"

    dates = month_data[month]
    for i, date_str in enumerate(dates, start=2):
        d_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        is_hol = is_legal_holiday(date_str, holidays, workdays)
        dur = get_training_duration(d_obj, is_hol)

        sheet.cell(row=i, column=1, value=date_str)
        sheet.cell(row=i, column=2, value="是" if is_hol else "否")
        sheet.cell(row=i, column=3, value=dur)

        # ✅ 按你给的公式写入金额（行号随 i 变化）
        sheet.cell(row=i, column=4, value=f'=IF(B{i}="","",IF(B{i}="是",C{i}*200,IF(B{i}="否",C{i}*100,"")))' )

    style_header(sheet)
    autofit_column_width(sheet)

    # 封面公式（仍按工时汇总计算金额）
    row = month + 1
    cover.cell(row=row, column=1, value=f"{month}月")
    cover.cell(row=row, column=2, value=f"=SUMIFS('{month}月'!C:C,'{month}月'!B:B,\"否\")")
    cover.cell(row=row, column=3, value=f"=SUMIFS('{month}月'!C:C,'{month}月'!B:B,\"是\")")
    cover.cell(row=row, column=4, value=f"=B{row}*100")
    cover.cell(row=row, column=5, value=f"=C{row}*200")
    cover.cell(row=row, column=6, value=f"=D{row}+E{row}")

# 合计行
total_row = 14
cover.cell(row=total_row, column=1, value="合计")
cover.cell(row=total_row, column=2, value="=SUM(B2:B13)")
cover.cell(row=total_row, column=3, value="=SUM(C2:C13)")
cover.cell(row=total_row, column=4, value="=SUM(D2:D13)")
cover.cell(row=total_row, column=5, value="=SUM(E2:E13)")
cover.cell(row=total_row, column=6, value="=SUM(F2:F13)")

style_header(cover)
autofit_column_width(cover)

output_path = os.path.join(folder_path, f"{TARGET_YEAR}全年训练日期汇总.xlsx")
wb.save(output_path)
print(f"\n✅ 生成完成：{output_path}")